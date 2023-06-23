# coding: utf-8 -*-
import streamlit as st
import pandas as pd
import openpyxl as excel
import numpy as np
import datetime as tempo

st.set_option('deprecation.showPyplotGlobalUse', False)
st.markdown("""
    <style>
        *, p, div {
            font-size: 22px;
            color: black;
        }
    </style>
""",unsafe_allow_html=True)

arquivo = excel.load_workbook("base_horas_extras_alpargatas.xlsx")
placas = ["OWF0F93 / GHW3B18","OWF0F93 / NNR9083","HKE0E36 / QFC7729","HKE0E36 / FBY3D91","GZV9A04 / OFY2053","GZV9A04 / EVO0I65","JUU2J75 / IZB9F00","KSW7626 / GMC7F93"]

tabAcidionar, tabBaixar, tabRemover = st.tabs(["Registrar","Baixar relatório","Remover linha"])

with tabAcidionar:
    opcoesFabrica = st.selectbox(
        label="Selecione a Fábrica:",
        options=["CARPINA","SANTA RITA"]
    )
    opcoesPlaca = st.selectbox(
        label="Selecione a placa:",
        options=placas
    )
    qtdViagens = st.number_input(
        label="Digite a quantidade de viagens:",
        min_value=0,
        max_value=10000
    )
    qtdCargaBatida = st.number_input(
        label="Digite a quantidade de cargas batidas:",
        min_value=0,
        max_value=10000
    )
    pagarAgregado = st.number_input(
        label="Digite o valor a pagar ao agregado:",
        min_value=0.0,
        max_value=100000.0
    )
    qtdDiariasPagar = st.number_input(
        label="Digite a quantidade de diárias a pagar:",
        min_value=0,
        max_value=100
    )
    valorCobrar = st.number_input(
        label="Digite o valor a cobrar da Alpargatas:",
        min_value=0.0,
        max_value=100000.0
    )
    qtdDiariasReceber = st.number_input(
        label="Digite a quantidade de diárias a receber:",
        min_value=0,
        max_value=100
    )
    inserirDados = st.button(
        label="Inserir"
    )
    if inserirDados:
        qtdLinhas = arquivo["baseApp"].max_row + 1
        arquivo["baseApp"].cell(row=qtdLinhas,column=1).value = opcoesFabrica
        arquivo["baseApp"].cell(row=qtdLinhas,column=2).value = opcoesPlaca
        arquivo["baseApp"].cell(row=qtdLinhas,column=3).value = qtdViagens
        arquivo["baseApp"].cell(row=qtdLinhas,column=4).value = qtdCargaBatida
        arquivo["baseApp"].cell(row=qtdLinhas,column=5).value = valorCobrar
        arquivo["baseApp"].cell(row=qtdLinhas,column=6).value = qtdDiariasPagar
        arquivo["baseApp"].cell(row=qtdLinhas,column=7).value = valorCobrar
        arquivo["baseApp"].cell(row=qtdLinhas,column=8).value = qtdDiariasReceber
        dia = tempo.datetime.today().day
        mes = tempo.datetime.today().month
        ano = tempo.datetime.today().year
        arquivo["baseApp"].cell(row=qtdLinhas,column=9).value = "{}/{}/{}".format(dia,mes,ano)
        try:
            arquivo.save("base_horas_extras_alpargatas.xlsx")
            st.success("Dados inseridos com sucesso")
        except:
            st.warning("Deu merda")

with tabBaixar:
    paraDownload = pd.DataFrame(
        pd.read_excel(
            "base_horas_extras_alpargatas.xlsx",
            sheet_name="baseApp"
        )
    ).to_csv(sep=";",index=False)
    st.dataframe(pd.DataFrame(pd.read_excel("base_horas_extras_alpargatas.xlsx")))
    st.download_button(
        label="Baixar relatório",
        data=paraDownload,
        file_name="horas_extra_alpargatas.csv"
    )

with tabRemover:
    st.dataframe(pd.DataFrame(pd.read_excel("base_horas_extras_alpargatas.xlsx")))
    removerLinha = st.number_input(
        label="Digite o índice para remover a linha:",
        min_value=0,
        max_value=10000
    )
    botaoRemover = st.button(
        label="Remover linha"
    )
    if botaoRemover:
        try:
            arquivo["baseApp"].delete_rows(idx=(removerLinha+2))
            arquivo.save("base_horas_extras_alpargatas.xlsx")
            st.success("Linha removida com sucesso")
        except:
            st.warning("Deu merda")